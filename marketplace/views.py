# marketplace/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import jdatetime
from .models import (
    MarketplaceSale, MarketplacePurchase, MarketplacePurchaseDetail, 
    DeliveryAddress, ProductOffer
)


@staff_member_required
def download_purchases_excel(request, sale_id):
    """دانلود خریدهای یک فروش به فرمت اکسل"""
    sale = get_object_or_404(MarketplaceSale, id=sale_id)
    
    # ایجاد فایل اکسل
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"خریدهای فروش {sale.product_offer.offer_id}"
    
    # تنظیمات استایل
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # هدرها
    headers = [
        'شناسه خرید', 'وزن خرید', 'تاریخ خرید', 'نام خریدار',
        'شماره همراه خریدار', 'شماره ملی خریدار', 'مبلغ پرداختی', 'نوع خرید'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # داده‌های خریدها
    purchases = sale.purchases.all().order_by('purchase_date')
    for row_idx, purchase in enumerate(purchases, 2):
        sheet.cell(row=row_idx, column=1, value=purchase.purchase_id)
        sheet.cell(row=row_idx, column=2, value=float(purchase.purchase_weight))
        
        # تبدیل تاریخ به شمسی
        try:
            jalali_date = jdatetime.date.fromgregorian(
                year=purchase.purchase_date.year,
                month=purchase.purchase_date.month,
                day=purchase.purchase_date.day
            )
            sheet.cell(row=row_idx, column=3, value=jalali_date.strftime("%Y/%m/%d"))
        except:
            sheet.cell(row=row_idx, column=3, value=purchase.purchase_date.strftime("%Y-%m-%d"))
        
        sheet.cell(row=row_idx, column=4, value=purchase.buyer_name)
        sheet.cell(row=row_idx, column=5, value=purchase.buyer_mobile)
        sheet.cell(row=row_idx, column=6, value=purchase.buyer_national_id)
        sheet.cell(row=row_idx, column=7, value=int(purchase.paid_amount))
        sheet.cell(row=row_idx, column=8, value=purchase.get_purchase_type_display())
    
    # تنظیم عرض ستون‌ها
    column_widths = [15, 12, 12, 25, 15, 12, 15, 12]
    for col_idx, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # ایجاد response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="purchases_{sale.product_offer.offer_id}.xlsx"'
    
    # نوشتن فایل در response
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response.write(virtual_workbook.getvalue())
    
    return response


@staff_member_required
def download_purchases_template(request):
    """دانلود نمونه فایل اکسل برای آپلود خریدها"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "نمونه خریدها"
    
    # تنظیمات استایل
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # هدرها
    headers = [
        'شناسه خرید', 'وزن خرید', 'تاریخ خرید', 'نام خریدار',
        'شماره همراه خریدار', 'شماره ملی خریدار', 'مبلغ پرداختی', 'نوع خرید'
    ]
    descriptions = [
        'شناسه یکتای خرید از بازارگاه',
        'وزن به تن (مثال: 12.5)',
        'تاریخ شمسی (1403/10/15)',
        'نام کامل خریدار',
        'شماره همراه (09xxxxxxxxx)',
        'کد ملی 10 رقمی',
        'مبلغ به ریال (عدد)',
        'نقدی یا توافقی'
    ]
    
    for col_idx, (header, desc) in enumerate(zip(headers, descriptions), 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        
        # توضیحات در ردیف دوم
        desc_cell = sheet.cell(row=2, column=col_idx, value=desc)
        desc_cell.font = Font(italic=True, color="666666")
    
    # نمونه داده
    sample_data = [
        ['BUY001', 25.5, '1403/10/15', 'احمد محمدی', '09123456789', '1234567890', 50000000, 'نقدی'],
        ['BUY002', 10.0, '1403/10/16', 'فاطمه احمدی', '09987654321', '0987654321', 20000000, 'توافقی'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # تنظیم عرض ستون‌ها
    for col_idx in range(1, 9):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
    
    # ایجاد response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="purchases_template.xlsx"'
    
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response.write(virtual_workbook.getvalue())
    
    return response


@staff_member_required
def upload_purchases_excel(request, sale_id):
    """آپلود فایل اکسل خریدها"""
    sale = get_object_or_404(MarketplaceSale, id=sale_id)
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            success_count = 0
            error_rows = []
            
            # خواندن داده‌ها از ردیف سوم (دو ردیف اول هدر و توضیحات)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                try:
                    if not any(row):
                        continue
                    
                    purchase_id = str(row[0]).strip() if row[0] else None
                    purchase_weight = float(row[1]) if row[1] else None
                    purchase_date_str = str(row[2]).strip() if row[2] else None
                    buyer_name = str(row[3]).strip() if row[3] else None
                    buyer_mobile = str(row[4]).strip() if row[4] else None
                    buyer_national_id = str(row[5]).strip() if row[5] else None
                    paid_amount = float(row[6]) if row[6] else None
                    purchase_type_str = str(row[7]).strip() if row[7] else None
                    
                    # بررسی فیلدهای اجباری
                    if not all([purchase_id, purchase_weight, purchase_date_str, buyer_name, buyer_mobile, buyer_national_id, paid_amount, purchase_type_str]):
                        error_rows.append(f"ردیف {row_idx}: داده‌های ناقص")
                        continue
                    
                    # اطمینان از صحت نوع داده‌ها
                    try:
                        purchase_weight = float(purchase_weight)
                        paid_amount = float(paid_amount)
                    except (ValueError, TypeError):
                        error_rows.append(f"ردیف {row_idx}: وزن یا مبلغ نامعتبر")
                        continue
                    
                    # تبدیل تاریخ شمسی به میلادی
                    try:
                        if '/' in purchase_date_str:
                            date_parts = purchase_date_str.split('/')
                            jalali_date = jdatetime.date(int(date_parts[0]), int(date_parts[1]), int(date_parts[2]))
                            purchase_date = jalali_date.togregorian()
                        else:
                            error_rows.append(f"ردیف {row_idx}: فرمت تاریخ نامعتبر")
                            continue
                    except:
                        error_rows.append(f"ردیف {row_idx}: تاریخ نامعتبر")
                        continue
                    
                    # تبدیل نوع خرید
                    purchase_type = 'cash' if purchase_type_str.lower() in ['نقدی', 'cash'] else 'agreement'
                    
                    # بررسی تکراری بودن شناسه خرید
                    if MarketplacePurchase.objects.filter(purchase_id=purchase_id).exists():
                        error_rows.append(f"ردیف {row_idx}: شناسه خرید تکراری ({purchase_id})")
                        continue
                    
                    # ایجاد خرید
                    purchase = MarketplacePurchase.objects.create(
                        marketplace_sale=sale,
                        purchase_id=purchase_id,
                        purchase_weight=purchase_weight,
                        purchase_date=purchase_date,
                        buyer_name=buyer_name,
                        buyer_mobile=buyer_mobile,
                        buyer_national_id=buyer_national_id,
                        paid_amount=paid_amount,
                        purchase_type=purchase_type
                    )
                    
                    # ایجاد جزئیات خرید
                    MarketplacePurchaseDetail.objects.get_or_create(
                        purchase=purchase
                    )
                    
                    success_count += 1
                    
                except ValueError as e:
                    error_rows.append(f"ردیف {row_idx}: مقدار عددی نامعتبر")
                except Exception as e:
                    error_rows.append(f"ردیف {row_idx}: {str(e)}")
            
            # نمایش نتایج
            if success_count > 0:
                messages.success(request, f'✅ {success_count} خرید با موفقیت اضافه شد')
            
            if error_rows:
                error_message = '❌ خطا در ردیف‌های: ' + ', '.join(error_rows[:5])
                if len(error_rows) > 5:
                    error_message += f' و {len(error_rows) - 5} خطای دیگر'
                messages.warning(request, error_message)
                
        except Exception as e:
            messages.error(request, f'❌ خطا در خواندن فایل: {str(e)}')
        
        return redirect(f'/admin/marketplace/marketplacesale/{sale_id}/change/')
    
    # نمایش فرم آپلود
    context = {
        'sale': sale,
        'title': f'آپلود خریدهای فروش {sale.product_offer.offer_id}'
    }
    return render(request, 'marketplace/upload_purchases.html', context)


@staff_member_required
def download_delivery_template(request):
    """دانلود نمونه فایل اکسل برای آدرس‌های تحویل"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "نمونه آدرس تحویل"
    
    # تنظیمات استایل
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # تمام فیلدهای مورد نیاز
    headers = [
        'کد', 'وزن کل خرید', 'تاریخ خرید', 'قیمت هر واحد', 'شماره پیگیری',
        'استان', 'شهرستان', 'مبلغ پرداختی', 'شماره حساب خریدار', 'کد کوتاژ',
        'عنوان کالا', 'توضیحات', 'شیوه پرداخت', 'شناسه عرضه', 'تاریخ ثبت آدرس',
        'شناسه تخصیص', 'نام خریدار', 'شناسه ملی خریدار', 'کدپستی خریدار',
        'آدرس خریدار', 'شناسه واریز', 'شماره همراه خریدار', 'شناسه یکتا خریدار',
        'نوع کاربری خریدار', 'نام تحویل گیرنده', 'شناسه یکتای تحویل', 'تک',
        'جفت', 'تریلی', 'آدرس تحویل', 'کد پستی تحویل', 'شماره هماهنگی تحویل',
        'کد ملی تحویل', 'وزن سفارش', 'بازه 1 پرداخت توافقی (روز)',
        'بازه 2 پرداخت توافقی (روز)', 'بازه 3 پرداخت توافقی (روز)',
        'مبلغ بازه 1 توافقی-ریال', 'مبلغ بازه 2 توافقی-ریال', 'مبلغ بازه 3 توافقی-ریال',
        'وزن بارنامه شده', 'وزن بارنامه نشده'
    ]
    
    # نوشتن هدرها
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # نمونه داده
    sample_data = [
        'BUY001', 25.5, '1403/10/15', 2000000, 'TRK001', 'تهران', 'تهران',
        51000000, '1234567890123456', 'CTG001', 'سیمان پرتلند', 'سفارش عادی',
        'نقدی', 'OFF001', '1403/10/16', 'ADDR001', 'احمد محمدی', '1234567890',
        '1234567890', 'تهران، خیابان آزادی', 'DEP001', '09123456789', 'USR001',
        'حقیقی', 'علی احمدی', 'REC001', 'بله', 'خیر', 'خیر',
        'تهران، خیابان انقلاب', '1234567890', '09987654321', '0987654321',
        25.5, '', '', '', '', '', '', 0, 0
    ]
    
    # نوشتن نمونه داده
    for col_idx, value in enumerate(sample_data, 1):
        sheet.cell(row=2, column=col_idx, value=value)
    
    # تنظیم عرض ستون‌ها
    for col_idx in range(1, len(headers) + 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
    
    # ایجاد response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="delivery_addresses_template.xlsx"'
    
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response.write(virtual_workbook.getvalue())
    
    return response


@staff_member_required
def upload_delivery_addresses(request, purchase_detail_id):
    """آپلود آدرس‌های تحویل با پردازش بهبود یافته"""
    purchase_detail = get_object_or_404(MarketplacePurchaseDetail, id=purchase_detail_id)
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            success_count = 0
            error_rows = []
            
            # حذف آدرس‌های قبلی
            old_count = purchase_detail.delivery_addresses.count()
            purchase_detail.delivery_addresses.all().delete()
            
            if old_count > 0:
                messages.info(request, f'🔄 {old_count} آدرس قبلی حذف شد')
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # ردیف خالی
                    continue
                    
                try:
                    # اعتبارسنجی فیلدهای اصلی
                    code = str(row[0]).strip() if row[0] else ''
                    if not code:
                        error_rows.append(f"ردیف {row_idx}: کد خالی است")
                        continue
                    
                    # پردازش وزن کل خرید
                    try:
                        total_purchase_weight = float(row[1]) if row[1] else 0
                    except (ValueError, TypeError):
                        error_rows.append(f"ردیف {row_idx}: وزن نامعتبر")
                        continue
                    
                    # پردازش تاریخ خرید
                    try:
                        if isinstance(row[2], str):
                            date_parts = row[2].split('/')
                            if len(date_parts) == 3:
                                purchase_date = jdate(int(date_parts[0]), int(date_parts[1]), int(date_parts[2])).togregorian()
                            else:
                                purchase_date = purchase_detail.purchase.purchase_date
                        else:
                            purchase_date = purchase_detail.purchase.purchase_date
                    except:
                        purchase_date = purchase_detail.purchase.purchase_date
                    
                    # پردازش تاریخ ثبت آدرس
                    try:
                        if isinstance(row[14], str) and row[14]:
                            date_parts = row[14].split('/')
                            if len(date_parts) == 3:
                                address_registration_date = jdate(int(date_parts[0]), int(date_parts[1]), int(date_parts[2])).togregorian()
                            else:
                                address_registration_date = purchase_detail.purchase.purchase_date
                        else:
                            address_registration_date = purchase_detail.purchase.purchase_date
                    except:
                        address_registration_date = purchase_detail.purchase.purchase_date
                    
                    # تابع کمکی برای پاکسازی رشته‌ها
                    def clean_string(value, max_length=None):
                        if value is None:
                            return ''
                        result = str(value).strip()
                        if max_length and len(result) > max_length:
                            result = result[:max_length]
                        return result
                    
                    # تابع کمکی برای پاکسازی اعداد
                    def clean_numeric_string(value, max_length=20):
                        if value is None:
                            return ''
                        # حذف همه کاراکترهای غیرعددی
                        result = ''.join(filter(str.isdigit, str(value)))
                        if len(result) > max_length:
                            result = result[:max_length]
                        return result
                    
                    # پردازش فیلدهای عددی
                    def safe_float(value, default=0):
                        try:
                            return float(value) if value else default
                        except (ValueError, TypeError):
                            return default
                    
                    def safe_int(value, default=None):
                        try:
                            return int(value) if value else default
                        except (ValueError, TypeError):
                            return default
                    
                    # ایجاد آدرس تحویل با پاکسازی داده‌ها
                    delivery_address = DeliveryAddress.objects.create(
                        purchase_detail=purchase_detail,
                        code=clean_string(code, 100),
                        total_purchase_weight=total_purchase_weight,
                        purchase_date=purchase_date,
                        unit_price=safe_float(row[3]),
                        tracking_number=clean_string(row[4], 100),
                        province=clean_string(row[5], 100),
                        city=clean_string(row[6], 100),
                        paid_amount=safe_float(row[7]),
                        buyer_account_number=clean_string(row[8], 50),
                        cottage_code=clean_string(row[9], 50),
                        product_title=clean_string(row[10], 300),
                        description=clean_string(row[11]),
                        payment_method=clean_string(row[12], 50),
                        offer_id=clean_string(row[13], 100),
                        address_registration_date=address_registration_date,
                        assignment_id=clean_string(row[15], 100) or f'ADDR_{code}_{row_idx}',
                        buyer_name=clean_string(row[16], 200),
                        
                        # فیلدهای با طول افزایش یافته - FIX برای خطای character varying(10)
                        buyer_national_id=clean_numeric_string(row[17], 20),
                        buyer_postal_code=clean_numeric_string(row[18], 20),
                        
                        buyer_address=clean_string(row[19]),
                        deposit_id=clean_string(row[20], 100),
                        
                        # شماره تلفن‌ها با طول افزایش یافته
                        buyer_mobile=clean_numeric_string(row[21], 20),
                        
                        buyer_unique_id=clean_string(row[22], 100),
                        buyer_user_type='individual' if str(row[23]).strip() == 'حقیقی' else 'company',
                        recipient_name=clean_string(row[24], 200),
                        recipient_unique_id=clean_string(row[25], 100),
                        
                        # وسایل حمل
                        vehicle_single=str(row[26]).strip().lower() in ['بله', 'true', '1'] if row[26] else False,
                        vehicle_double=str(row[27]).strip().lower() in ['بله', 'true', '1'] if row[27] else False,
                        vehicle_trailer=str(row[28]).strip().lower() in ['بله', 'true', '1'] if row[28] else False,
                        
                        # آدرس تحویل با فیلدهای طول افزایش یافته
                        delivery_address=clean_string(row[29]),
                        delivery_postal_code=clean_numeric_string(row[30], 20),
                        coordination_phone=clean_numeric_string(row[31], 20),
                        delivery_national_id=clean_numeric_string(row[32], 20),
                        
                        order_weight=safe_float(row[33]),
                        
                        # بازه‌های پرداخت توافقی
                        payment_period_1_days=safe_int(row[34]),
                        payment_period_2_days=safe_int(row[35]),
                        payment_period_3_days=safe_int(row[36]),
                        payment_amount_1=safe_float(row[37]),
                        payment_amount_2=safe_float(row[38]),
                        payment_amount_3=safe_float(row[39]),
                        
                        # وزن‌های بارنامه
                        shipped_weight=safe_float(row[40]),
                        unshipped_weight=safe_float(row[41]),
                    )
                    
                    success_count += 1
                    
                except ValueError as e:
                    error_rows.append(f"ردیف {row_idx}: مقدار عددی نامعتبر - {str(e)}")
                except Exception as e:
                    error_rows.append(f"ردیف {row_idx}: {str(e)}")
            
            # نمایش نتایج
            if success_count > 0:
                messages.success(request, f'✅ {success_count} آدرس تحویل با موفقیت اضافه شد')
            
            if error_rows:
                error_message = '❌ خطا در ردیف‌های: ' + ', '.join(error_rows[:5])
                if len(error_rows) > 5:
                    error_message += f' و {len(error_rows) - 5} خطای دیگر'
                messages.warning(request, error_message)
                
        except Exception as e:
            messages.error(request, f'❌ خطا در خواندن فایل: {str(e)}')
        
        return redirect(f'/admin/marketplace/marketplacepurchasedetail/{purchase_detail_id}/change/')
    
    # نمایش فرم آپلود
    context = {
        'purchase_detail': purchase_detail,
        'title': f'آپلود آدرس‌های تحویل برای خرید {purchase_detail.purchase.purchase_id}'
    }
    return render(request, 'marketplace/upload_delivery_addresses.html', context)