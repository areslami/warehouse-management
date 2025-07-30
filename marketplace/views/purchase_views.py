# marketplace/views/purchase_views.py
from django.shortcuts import get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import jdatetime
from decimal import Decimal
from ..models import MarketplaceSale, MarketplacePurchase, MarketplacePurchaseDetail
from .mixins import ExcelResponseMixin, PersianDateMixin, DataCleaningMixin, HeaderBasedExcelMixin


class PurchaseExcelHandler(ExcelResponseMixin, PersianDateMixin, DataCleaningMixin, HeaderBasedExcelMixin):
    """Handler for purchase-related Excel operations"""
    
    def get_field_mapping(self):
        """Return mapping of field names to possible Excel header variations"""
        return {
            'purchase_id': ['شناسه خرید', 'کد خرید', 'شناسه', 'ID'],
            'cottage_number': ['شماره کوتاژ', 'کوتاژ', 'شماره کوتاژ (با شماره کوتاژ مربوط به عرضه کنترل شود. در صورت عدم تطابق بعد از آپلود و قبل از ثبت در سیستم اخطار بدهد)'],
            'description': ['توضیحات', 'شرح', 'Description'],
            'purchase_weight': ['وزن خرید شده-Kg', 'وزن خرید', 'وزن', 'Weight'],
            'province': ['استان', 'Province'],
            'purchase_date': ['تاریخ خرید', 'تاریخ', 'Date'],
            'paid_amount': ['مبلغ پرداختی-ریال', 'مبلغ پرداختی', 'مبلغ', 'Amount'],
            'unit_price': ['قیمت هر واحد-ریال', 'قیمت واحد', 'قیمت', 'Price'],
            'delivery_date': ['تاریخ تحویل', 'Delivery Date'],
            'tracking_number': ['شماره پیگیری', 'پیگیری', 'Tracking'],
            'document_registration_date': ['تاریخ ثبت سند', 'تاریخ سند'],
            'product_title': ['عنوان کالا', 'کالا', 'محصول', 'Product'],
            'buyer_national_id': ['کد ملی خریدار', 'کد ملی', 'شناسه ملی', 'National ID'],
            'buyer_account_number': ['شماره حساب خریدار', 'شماره حساب', 'Account'],
            'buyer_mobile': ['شماره همراه خریدار', 'شماره همراه', 'موبایل', 'Mobile'],
            'buyer_name': ['نام خریدار', 'نام', 'Name'],
            'purchase_type': ['شیوه پرداخت', 'نوع پرداخت', 'Payment Type'],
            'agreement_period_1': ['بازه 1 پرداخت توافقی (روز)', 'بازه 1'],
            'agreement_period_2': ['بازه 2 پرداخت توافقی (روز)', 'بازه 2'],
            'agreement_period_3': ['بازه 3 پرداخت توافقی (روز)', 'بازه 3'],
            'agreement_amount_1': ['مبلغ بازه 1 توافقی-ریال', 'مبلغ بازه 1'],
            'agreement_amount_2': ['مبلغ بازه 2 توافقی-ریال', 'مبلغ بازه 2'],
            'agreement_amount_3': ['مبلغ بازه 3 توافقی-ریال', 'مبلغ بازه 3'],
            'supply_id': ['شناسه عرضه', 'عرضه', 'Supply ID']
        }
    
    def create_purchases_excel(self, sale):
        """Create Excel file with purchase data"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"خریدهای فروش {sale.product_offer.offer_id}"
        
        # Headers
        headers = [
            'شناسه خرید', 'وزن خرید', 'تاریخ خرید', 'نام خریدار',
            'شماره همراه خریدار', 'شماره ملی خریدار', 'مبلغ پرداختی', 'نوع خرید'
        ]
        
        # Create styled header
        self.create_styled_header(sheet, headers)
        
        # Purchase data
        purchases = sale.purchases.all().order_by('purchase_date')
        for row_idx, purchase in enumerate(purchases, 2):
            sheet.cell(row=row_idx, column=1, value=purchase.purchase_id)
            sheet.cell(row=row_idx, column=2, value=float(purchase.purchase_weight))
            sheet.cell(row=row_idx, column=3, value=self.gregorian_to_persian(purchase.purchase_date))
            sheet.cell(row=row_idx, column=4, value=purchase.buyer_name)
            sheet.cell(row=row_idx, column=5, value=purchase.buyer_mobile)
            sheet.cell(row=row_idx, column=6, value=purchase.buyer_national_id)
            sheet.cell(row=row_idx, column=7, value=int(purchase.paid_amount))
            sheet.cell(row=row_idx, column=8, value=purchase.get_purchase_type_display())
        
        # Set column widths
        column_widths = [15, 12, 12, 25, 15, 12, 15, 12]
        for col_idx, width in enumerate(column_widths, 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        return workbook
    
    def create_purchases_template(self):
        """Create Excel template for purchase uploads"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "نمونه خریدها"
        
        # Headers with all required fields
        headers = [
            'شناسه خرید', 'شماره کوتاژ', 'توضیحات', 'وزن خرید شده-Kg', 'استان',
            'تاریخ خرید', 'مبلغ پرداختی-ریال', 'قیمت هر واحد-ریال', 'تاریخ تحویل',
            'شماره پیگیری', 'تاریخ ثبت سند', 'عنوان کالا', 'کد ملی خریدار',
            'شماره حساب خریدار', 'شماره همراه خریدار', 'نام خریدار', 'شیوه پرداخت',
            'بازه 1 پرداخت توافقی (روز)', 'بازه 2 پرداخت توافقی (روز)', 'بازه 3 پرداخت توافقی (روز)',
            'مبلغ بازه 1 توافقی-ریال', 'مبلغ بازه 2 توافقی-ریال', 'مبلغ بازه 3 توافقی-ریال', 'شناسه عرضه'
        ]
        
        # Create styled header
        self.create_styled_header(sheet, headers)
        
        # Example data row
        example_data = [
            'PUR001', 'COT123', 'نمونه توضیحات', '2500', 'تهران',
            '1403/05/15', '50000000', '20000', '1403/05/20',
            'TRK001', '1403/05/16', 'برنج درجه یک', '1234567890',
            '6037997123456789', '09123456789', 'احمد محمدی', 'نقدی',
            '', '', '', '', '', '', 'OFF001'
        ]
        
        for col_idx, data in enumerate(example_data, 1):
            cell = sheet.cell(row=2, column=col_idx, value=data)
            cell.alignment = Alignment(horizontal="center")
        
        # Set column widths
        column_widths = [12, 12, 20, 15, 12, 12, 15, 15, 12,
                        12, 12, 20, 15, 20, 15, 20, 12,
                        10, 10, 10, 15, 15, 15, 12]
        for col_idx, width in enumerate(column_widths, 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        return workbook
    
    def process_purchases_upload(self, file, sale):
        """Process uploaded Excel file and create purchases"""
        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            
            # Get header mapping from first row
            header_mapping = self.get_header_mapping(sheet)
            field_mapping = self.get_field_mapping()
            
            # Validate required headers
            required_headers = [
                ['شناسه خرید', 'کد خرید', 'شناسه'],
                ['وزن خرید شده-Kg', 'وزن خرید', 'وزن'],
                ['تاریخ خرید', 'تاریخ'],
                ['نام خریدار', 'نام']
            ]
            
            missing_headers = self.validate_required_headers(header_mapping, required_headers)
            if missing_headers:
                return [], [f'هدرهای الزامی موجود نیست: {", ".join(missing_headers)}']
            
            created_purchases = []
            errors = []
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue
                
                try:
                    # Extract data using header-based mapping
                    row_data = self.extract_row_data(row, header_mapping, field_mapping)
                    
                    # Process extracted data
                    purchase_id = str(row_data['purchase_id']).strip() if row_data['purchase_id'] else None
                    cottage_number = str(row_data['cottage_number']).strip() if row_data['cottage_number'] else ''
                    description = str(row_data['description']).strip() if row_data['description'] else ''
                    purchase_weight = self.clean_numeric_string(row_data['purchase_weight'])
                    province = str(row_data['province']).strip() if row_data['province'] else ''
                    purchase_date_str = str(row_data['purchase_date']).strip() if row_data['purchase_date'] else None
                    paid_amount = self.clean_numeric_string(row_data['paid_amount'])
                    unit_price = self.clean_numeric_string(row_data['unit_price'])
                    delivery_date_str = str(row_data['delivery_date']).strip() if row_data['delivery_date'] else None
                    tracking_number = str(row_data['tracking_number']).strip() if row_data['tracking_number'] else ''
                    document_date_str = str(row_data['document_registration_date']).strip() if row_data['document_registration_date'] else None
                    product_title = str(row_data['product_title']).strip() if row_data['product_title'] else ''
                    buyer_national_id = self.clean_national_id(row_data['buyer_national_id'])
                    buyer_account_number = str(row_data['buyer_account_number']).strip() if row_data['buyer_account_number'] else ''
                    buyer_mobile = self.clean_phone_number(row_data['buyer_mobile'])
                    buyer_name = str(row_data['buyer_name']).strip() if row_data['buyer_name'] else ''
                    purchase_type_str = str(row_data['purchase_type']).strip() if row_data['purchase_type'] else ''
                    agreement_period_1 = self.clean_numeric_string(row_data['agreement_period_1']) if row_data['agreement_period_1'] else None
                    agreement_period_2 = self.clean_numeric_string(row_data['agreement_period_2']) if row_data['agreement_period_2'] else None
                    agreement_period_3 = self.clean_numeric_string(row_data['agreement_period_3']) if row_data['agreement_period_3'] else None
                    agreement_amount_1 = self.clean_numeric_string(row_data['agreement_amount_1']) if row_data['agreement_amount_1'] else None
                    agreement_amount_2 = self.clean_numeric_string(row_data['agreement_amount_2']) if row_data['agreement_amount_2'] else None
                    agreement_amount_3 = self.clean_numeric_string(row_data['agreement_amount_3']) if row_data['agreement_amount_3'] else None
                    supply_id = str(row_data['supply_id']).strip() if row_data['supply_id'] else ''
                    
                    # Validate required fields
                    if not purchase_id:
                        errors.append(f'ردیف {row_idx}: شناسه خرید الزامی است')
                        continue
                    
                    if not purchase_weight or purchase_weight <= 0:
                        errors.append(f'ردیف {row_idx}: وزن خرید باید بیشتر از صفر باشد')
                        continue
                    
                    # Convert Persian dates
                    purchase_date = self.persian_to_gregorian(purchase_date_str)
                    if not purchase_date:
                        errors.append(f'ردیف {row_idx}: تاریخ خرید نامعتبر است')
                        continue
                    
                    delivery_date = None
                    if delivery_date_str:
                        delivery_date = self.persian_to_gregorian(delivery_date_str)
                    
                    document_date = None
                    if document_date_str:
                        document_date = self.persian_to_gregorian(document_date_str)
                    
                    # Map purchase type
                    purchase_type_mapping = {
                        'نقدی': 'cash',
                        'توافقی': 'agreement',
                        'ترکیبی': 'mixed'
                    }
                    purchase_type = purchase_type_mapping.get(purchase_type_str, 'cash')
                    
                    # Check for duplicate purchase_id
                    if MarketplacePurchase.objects.filter(purchase_id=purchase_id).exists():
                        errors.append(f'ردیف {row_idx}: شناسه خرید {purchase_id} قبلاً ثبت شده است')
                        continue
                    
                    # Create purchase with all fields
                    purchase = MarketplacePurchase.objects.create(
                        marketplace_sale=sale,
                        purchase_id=purchase_id,
                        cottage_number=cottage_number,
                        description=description,
                        purchase_weight=Decimal(str(purchase_weight)),
                        province=province,
                        purchase_date=purchase_date,
                        paid_amount=Decimal(str(paid_amount)),
                        unit_price=Decimal(str(unit_price)),
                        delivery_date=delivery_date,
                        tracking_number=tracking_number,
                        document_registration_date=document_date,
                        product_title=product_title,
                        buyer_national_id=buyer_national_id,
                        buyer_account_number=buyer_account_number,
                        buyer_mobile=buyer_mobile,
                        buyer_name=buyer_name,
                        purchase_type=purchase_type,
                        agreement_period_1=int(agreement_period_1) if agreement_period_1 else None,
                        agreement_period_2=int(agreement_period_2) if agreement_period_2 else None,
                        agreement_period_3=int(agreement_period_3) if agreement_period_3 else None,
                        agreement_amount_1=Decimal(str(agreement_amount_1)) if agreement_amount_1 else None,
                        agreement_amount_2=Decimal(str(agreement_amount_2)) if agreement_amount_2 else None,
                        agreement_amount_3=Decimal(str(agreement_amount_3)) if agreement_amount_3 else None,
                        supply_id=supply_id
                    )
                    
                    # Create purchase detail
                    MarketplacePurchaseDetail.objects.create(purchase=purchase)
                    
                    created_purchases.append(purchase)
                    
                except Exception as e:
                    errors.append(f'ردیف {row_idx}: خطا در پردازش - {str(e)}')
            
            return created_purchases, errors
            
        except Exception as e:
            return [], [f'خطا در خواندن فایل: {str(e)}']


# View functions
@staff_member_required
def download_purchases_excel(request, sale_id):
    """دانلود خریدهای یک فروش به فرمت اکسل"""
    sale = get_object_or_404(MarketplaceSale, id=sale_id)
    handler = PurchaseExcelHandler()
    workbook = handler.create_purchases_excel(sale)
    filename = f"purchases_{sale.product_offer.offer_id}.xlsx"
    return handler.create_excel_response(workbook, filename)


@staff_member_required
def download_purchases_template(request):
    """دانلود نمونه فایل اکسل برای آپلود خریدها"""
    handler = PurchaseExcelHandler()
    workbook = handler.create_purchases_template()
    filename = "purchases_template.xlsx"
    return handler.create_excel_response(workbook, filename)


@staff_member_required
def upload_purchases_excel(request, sale_id):
    """آپلود خریدها از فایل اکسل"""
    sale = get_object_or_404(MarketplaceSale, id=sale_id)
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        handler = PurchaseExcelHandler()
        created_purchases, errors = handler.process_purchases_upload(request.FILES['excel_file'], sale)
        
        if errors:
            for error in errors:
                messages.error(request, error)
        
        if created_purchases:
            messages.success(
                request, 
                f'تعداد {len(created_purchases)} خرید با موفقیت ثبت شد'
            )
            
        # Redirect back to admin
        return redirect(f'/admin/marketplace/marketplacesale/{sale_id}/change/')
    
    # If not POST or no file, redirect back
    return redirect(f'/admin/marketplace/marketplacesale/{sale_id}/change/')