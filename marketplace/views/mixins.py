# marketplace/views/mixins.py
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import jdatetime
import io


class ExcelResponseMixin:
    """Mixin for creating Excel HTTP responses"""
    
    @staticmethod
    def create_excel_response(workbook, filename):
        """Create HTTP response with Excel file"""
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @staticmethod
    def create_styled_header(worksheet, headers, row=1):
        """Create styled header row in worksheet"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        return worksheet


class HeaderBasedExcelMixin:
    """Mixin for header-based Excel processing"""
    
    def get_header_mapping(self, sheet):
        """Extract header mapping from first row of Excel sheet"""
        headers = {}
        first_row = sheet[1]  # First row contains headers
        
        for idx, cell in enumerate(first_row, 1):
            if cell.value:
                header_name = str(cell.value).strip()
                headers[header_name] = idx
        
        return headers
    
    def get_field_mapping(self):
        """Return mapping of field names to possible Excel header variations
        Override this method in child classes"""
        return {}
    
    def extract_row_data(self, row, header_mapping, field_mapping):
        """Extract data from Excel row based on header mapping"""
        data = {}
        
        for field_name, possible_headers in field_mapping.items():
            value = None
            
            # Try each possible header variation
            for header in possible_headers:
                if header in header_mapping:
                    col_idx = header_mapping[header]
                    if col_idx <= len(row):
                        cell_value = row[col_idx - 1]  # Convert to 0-based index
                        if cell_value is not None and str(cell_value).strip():
                            value = cell_value
                            break
            
            data[field_name] = value
        
        return data
    
    def validate_required_headers(self, header_mapping, required_headers):
        """Validate that required headers exist in Excel file"""
        missing_headers = []
        
        for required_header_group in required_headers:
            # Each group can have multiple possible names
            if isinstance(required_header_group, list):
                found = any(header in header_mapping for header in required_header_group)
                if not found:
                    missing_headers.append(f"یکی از این هدرها: {', '.join(required_header_group)}")
            else:
                if required_header_group not in header_mapping:
                    missing_headers.append(required_header_group)
        
        return missing_headers


class PersianDateMixin:
    """Mixin for Persian date handling"""
    
    @staticmethod
    def persian_to_gregorian(persian_date_str):
        """Convert Persian date string to Gregorian date"""
        if not persian_date_str:
            return None
        try:
            # Parse Persian date (format: YYYY/MM/DD)
            parts = str(persian_date_str).split('/')
            if len(parts) == 3:
                jd = jdatetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
                return jd.togregorian()
        except (ValueError, TypeError):
            pass
        return None
    
    @staticmethod
    def gregorian_to_persian(gregorian_date):
        """Convert Gregorian date to Persian string"""
        if not gregorian_date:
            return ''
        try:
            jd = jdatetime.date.fromgregorian(date=gregorian_date)
            return jd.strftime('%Y/%m/%d')
        except (ValueError, TypeError):
            return str(gregorian_date)


class AdminRequiredMixin:
    """Mixin to require staff member access"""
    
    @method_decorator(staff_member_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


class DataCleaningMixin:
    """Mixin for common data cleaning operations"""
    
    @staticmethod
    def clean_numeric_string(value):
        """Clean and extract numeric value from string"""
        if value is None:
            return None
        try:
            # Remove non-numeric characters except decimal point
            cleaned = ''.join(c for c in str(value) if c.isdigit() or c == '.')
            return float(cleaned) if cleaned else 0
        except (ValueError, TypeError):
            return 0
    
    @staticmethod
    def clean_phone_number(phone):
        """Clean phone number to contain only digits"""
        if not phone:
            return ''
        return ''.join(filter(str.isdigit, str(phone)))
    
    @staticmethod
    def clean_national_id(national_id):
        """Clean national ID to contain only digits"""
        if not national_id:
            return ''
        return ''.join(filter(str.isdigit, str(national_id)))
    
    @staticmethod
    def clean_postal_code(postal_code):
        """Clean postal code to contain only digits"""
        if not postal_code:
            return ''
        return ''.join(filter(str.isdigit, str(postal_code)))