# marketplace/views/delivery_views.py
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
import openpyxl
from decimal import Decimal
from ..models import MarketplacePurchaseDetail, DeliveryAddress
from .mixins import ExcelResponseMixin, PersianDateMixin, DataCleaningMixin, HeaderBasedExcelMixin


class DeliveryAddressProcessor(ExcelResponseMixin, PersianDateMixin, DataCleaningMixin, HeaderBasedExcelMixin):
    """Processor for delivery address Excel operations"""
    
    def get_field_mapping(self):
        """Return mapping of field names to possible Excel header variations"""
        return {
            'code': ['کد', 'Code'],
            'total_purchase_weight': ['وزن کل خرید', 'وزن کل', 'Total Weight'],
            'purchase_date': ['تاریخ خرید', 'تاریخ', 'Date'],
            'unit_price': ['قیمت هر واحد', 'قیمت', 'Price'],
            'tracking_number': ['شماره پیگیری', 'پیگیری', 'Tracking'],
            'province': ['استان', 'Province'],
            'city': ['شهرستان', 'شهر', 'City'],
            'paid_amount': ['مبلغ پرداختی', 'مبلغ', 'Amount'],
            'buyer_account_number': ['شماره حساب خریدار', 'شماره حساب', 'Account'],
            'cottage_code': ['کد کوتاژ', 'کوتاژ', 'Cottage'],
            'product_title': ['عنوان کالا', 'کالا', 'Product'],
            'description': ['توضیحات', 'شرح', 'Description'],
            'payment_method': ['شیوه پرداخت', 'پرداخت', 'Payment'],
            'offer_id': ['شناسه عرضه', 'عرضه', 'Offer'],
            'address_registration_date': ['تاریخ ثبت آدرس', 'تاریخ ثبت'],
            'assignment_id': ['شناسه تخصیص', 'تخصیص', 'Assignment'],
            'buyer_name': ['نام خریدار', 'نام', 'Name'],
            'buyer_national_id': ['شناسه ملی خریدار', 'کد ملی', 'National ID'],
            'buyer_postal_code': ['کدپستی خریدار', 'کد پستی', 'Postal Code'],
            'buyer_address': ['آدرس خریدار', 'آدرس', 'Address'],
            'deposit_id': ['شناسه واریز', 'واریز', 'Deposit'],
            'buyer_mobile': ['شماره همراه خریدار', 'موبایل', 'Mobile'],
            'buyer_unique_id': ['شناسه یکتا خریدار', 'شناسه یکتا', 'Unique ID'],
            'buyer_user_type': ['نوع کاربری خریدار', 'نوع کاربر', 'User Type'],
            'recipient_name': ['نام تحویل گیرنده', 'تحویل گیرنده', 'Recipient'],
            'recipient_unique_id': ['شناسه یکتای تحویل', 'شناسه تحویل'],
            'vehicle_single': ['تک', 'Single'],
            'vehicle_double': ['جفت', 'Double'],
            'vehicle_trailer': ['تریلی', 'Trailer'],
            'delivery_address': ['آدرس تحویل', 'آدرس'],
            'delivery_postal_code': ['کد پستی تحویل', 'کد پستی'],
            'coordination_phone': ['شماره هماهنگی تحویل', 'هماهنگی'],
            'delivery_national_id': ['کد ملی تحویل', 'کد ملی'],
            'order_weight': ['وزن سفارش', 'وزن', 'Weight'],
            'payment_period_1_days': ['بازه 1 پرداخت توافقی (روز)', 'بازه 1'],
            'payment_period_2_days': ['بازه 2 پرداخت توافقی (روز)', 'بازه 2'],
            'payment_period_3_days': ['بازه 3 پرداخت توافقی (روز)', 'بازه 3'],
            'payment_amount_1': ['مبلغ بازه 1 توافقی-ریال', 'مبلغ بازه 1'],
            'payment_amount_2': ['مبلغ بازه 2 توافقی-ریال', 'مبلغ بازه 2'],
            'payment_amount_3': ['مبلغ بازه 3 توافقی-ریال', 'مبلغ بازه 3'],
            'shipped_weight': ['وزن بارنامه شده', 'بارنامه شده'],
            'unshipped_weight': ['وزن بارنامه نشده', 'بارنامه نشده']
        }
    
    def create_delivery_template(self):
        """Create Excel template for delivery addresses"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "نمونه آدرس تحویل"
        
        # Headers for all required fields
        headers = [
            'کد', 'وزن کل خرید', 'تاریخ خرید', 'قیمت هر واحد', 'شماره پیگیری',
            'استان', 'شهرستان', 'مبلغ پرداختی', 'شماره حساب خریدار', 'کد کوتاژ',
            'عنوان کالا', 'توضیحات', 'شیوه پرداخت', 'شناسه عرضه', 'تاریخ ثبت آدرس',
            'شناسه تخصیص', 'نام خریدار', 'شناسه ملی خریدار', 'کدپستی خریدار',
            'آدرس خریدار', 'شناسه واریز', 'شماره همراه خریدار', 'شناسه یکتا خریدار',
            'نوع کاربری خریدار', 'نام تحویل گیرنده', 'شناسه یکتای تحویل', 'تک',
            'جفت', 'تریلی', 'آدرس تحویل', 'کد پستی تحویل', 'شماره هماهنگی تحویل',
            'کد ملی تحویل', 'وزن سفارش', 'بازه 1 پرداخت توافقی (روز)',
            'بازه 2 پرداخت توافقی (روز)', 'بازه 3 پرداخت توافقی (روز)',
            'مبلغ بازه 1 توافقی-ریال', 'مبلغ بازه 2 توافقی-ریال', 'مبلغ بازه 3 توافقی-ریال',
            'وزن بارنامه شده', 'وزن بارنامه نشده'
        ]
        
        # Create styled header
        self.create_styled_header(sheet, headers)
        
        # Sample data
        sample_data = [
            'BUY001', 25.5, '1403/10/15', 2000000, 'TRK001', 'تهران', 'تهران',
            51000000, '1234567890123456', 'CTG001', 'سیمان پرتلند', 'سفارش عادی',
            'نقدی', 'OFF001', '1403/10/16', 'ADDR001', 'احمد محمدی', '1234567890',
            '1234567890', 'تهران، خیابان آزادی', 'DEP001', '09123456789', 'USR001',
            'حقیقی', 'علی احمدی', 'REC001', 'بله', 'خیر', 'خیر',
            'تهران، خیابان انقلاب', '1234567890', '09987654321', '0987654321',
            25.5, '', '', '', '', '', '', 0, 0
        ]
        
        # Write sample data
        for col_idx, value in enumerate(sample_data, 1):
            sheet.cell(row=2, column=col_idx, value=value)
        
        # Set column widths
        for col_idx in range(1, len(headers) + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
        
        return workbook
    
    def process_delivery_upload(self, file, purchase_detail):
        """Process uploaded delivery addresses Excel file"""
        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            
            # Get header mapping from first row
            header_mapping = self.get_header_mapping(sheet)
            field_mapping = self.get_field_mapping()
            
            # Validate required headers
            required_headers = [
                ['کد', 'Code'],
                ['شناسه تخصیص', 'تخصیص'],
                ['نام خریدار', 'نام']
            ]
            
            missing_headers = self.validate_required_headers(header_mapping, required_headers)
            if missing_headers:
                return 0, [f'هدرهای الزامی موجود نیست: {", ".join(missing_headers)}'], 0
            
            success_count = 0
            error_rows = []
            
            # Remove old addresses
            old_count = purchase_detail.delivery_addresses.count()
            purchase_detail.delivery_addresses.all().delete()
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue
                
                try:
                    # Extract data using header-based mapping
                    row_data = self.extract_row_data(row, header_mapping, field_mapping)
                    
                    # Validate main fields
                    code = str(row_data['code']).strip() if row_data['code'] else ''
                    if not code:
                        error_rows.append(f"ردیف {row_idx}: کد خالی است")
                        continue
                    
                    # Process fields with data cleaning
                    total_purchase_weight = self._safe_decimal(row_data['total_purchase_weight'])
                    purchase_date = self.persian_to_gregorian(str(row_data['purchase_date']).strip() if row_data['purchase_date'] else '')
                    unit_price = self._safe_decimal(row_data['unit_price'])
                    tracking_number = str(row_data['tracking_number']).strip() if row_data['tracking_number'] else ''
                    province = str(row_data['province']).strip() if row_data['province'] else ''
                    city = str(row_data['city']).strip() if row_data['city'] else ''
                    paid_amount = self._safe_decimal(row_data['paid_amount'])
                    buyer_account_number = str(row_data['buyer_account_number']).strip() if row_data['buyer_account_number'] else ''
                    cottage_code = str(row_data['cottage_code']).strip() if row_data['cottage_code'] else ''
                    product_title = str(row_data['product_title']).strip() if row_data['product_title'] else ''
                    description = str(row_data['description']).strip() if row_data['description'] else ''
                    payment_method = str(row_data['payment_method']).strip() if row_data['payment_method'] else ''
                    offer_id = str(row_data['offer_id']).strip() if row_data['offer_id'] else ''
                    address_registration_date = self.persian_to_gregorian(str(row_data['address_registration_date']).strip() if row_data['address_registration_date'] else '')
                    assignment_id = str(row_data['assignment_id']).strip() if row_data['assignment_id'] else ''
                    buyer_name = str(row_data['buyer_name']).strip() if row_data['buyer_name'] else ''
                    buyer_national_id = self.clean_national_id(row_data['buyer_national_id'])
                    buyer_postal_code = self.clean_postal_code(row_data['buyer_postal_code'])
                    buyer_address = str(row_data['buyer_address']).strip() if row_data['buyer_address'] else ''
                    deposit_id = str(row_data['deposit_id']).strip() if row_data['deposit_id'] else ''
                    buyer_mobile = self.clean_phone_number(row_data['buyer_mobile'])
                    buyer_unique_id = str(row_data['buyer_unique_id']).strip() if row_data['buyer_unique_id'] else ''
                    buyer_user_type = self._map_user_type(str(row_data['buyer_user_type']).strip() if row_data['buyer_user_type'] else '')
                    recipient_name = str(row_data['recipient_name']).strip() if row_data['recipient_name'] else ''
                    recipient_unique_id = str(row_data['recipient_unique_id']).strip() if row_data['recipient_unique_id'] else ''
                    
                    # Vehicle types
                    vehicle_single = self._safe_boolean(row_data['vehicle_single'])
                    vehicle_double = self._safe_boolean(row_data['vehicle_double'])
                    vehicle_trailer = self._safe_boolean(row_data['vehicle_trailer'])
                    
                    delivery_address = str(row_data['delivery_address']).strip() if row_data['delivery_address'] else ''
                    delivery_postal_code = self.clean_postal_code(row_data['delivery_postal_code'])
                    coordination_phone = self.clean_phone_number(row_data['coordination_phone'])
                    delivery_national_id = self.clean_national_id(row_data['delivery_national_id'])
                    order_weight = self._safe_decimal(row_data['order_weight'])
                    
                    # Payment periods (optional)
                    payment_period_1_days = self._safe_integer(row_data['payment_period_1_days'])
                    payment_period_2_days = self._safe_integer(row_data['payment_period_2_days'])
                    payment_period_3_days = self._safe_integer(row_data['payment_period_3_days'])
                    payment_amount_1 = self._safe_decimal(row_data['payment_amount_1'])
                    payment_amount_2 = self._safe_decimal(row_data['payment_amount_2'])
                    payment_amount_3 = self._safe_decimal(row_data['payment_amount_3'])
                    
                    # Shipping weights
                    shipped_weight = self._safe_decimal(row_data['shipped_weight'])
                    unshipped_weight = self._safe_decimal(row_data['unshipped_weight'])
                    
                    # Validate required fields
                    if not assignment_id:
                        error_rows.append(f"ردیف {row_idx}: شناسه تخصیص الزامی است")
                        continue
                    
                    if not buyer_name:
                        error_rows.append(f"ردیف {row_idx}: نام خریدار الزامی است")
                        continue
                    
                    # Check for duplicate assignment_id
                    if DeliveryAddress.objects.filter(assignment_id=assignment_id).exists():
                        error_rows.append(f"ردیف {row_idx}: شناسه تخصیص {assignment_id} قبلاً ثبت شده است")
                        continue
                    
                    # Create delivery address
                    DeliveryAddress.objects.create(
                        purchase_detail=purchase_detail,
                        code=code,
                        total_purchase_weight=total_purchase_weight,
                        purchase_date=purchase_date or purchase_detail.purchase.purchase_date,
                        unit_price=unit_price,
                        tracking_number=tracking_number,
                        province=province,
                        city=city,
                        paid_amount=paid_amount,
                        buyer_account_number=buyer_account_number,
                        cottage_code=cottage_code,
                        product_title=product_title,
                        description=description,
                        payment_method=payment_method,
                        offer_id=offer_id,
                        address_registration_date=address_registration_date or purchase_detail.purchase.purchase_date,
                        assignment_id=assignment_id,
                        buyer_name=buyer_name,
                        buyer_national_id=buyer_national_id,
                        buyer_postal_code=buyer_postal_code,
                        buyer_address=buyer_address,
                        deposit_id=deposit_id,
                        buyer_mobile=buyer_mobile,
                        buyer_unique_id=buyer_unique_id,
                        buyer_user_type=buyer_user_type,
                        recipient_name=recipient_name,
                        recipient_unique_id=recipient_unique_id,
                        vehicle_single=vehicle_single,
                        vehicle_double=vehicle_double,
                        vehicle_trailer=vehicle_trailer,
                        delivery_address=delivery_address,
                        delivery_postal_code=delivery_postal_code,
                        coordination_phone=coordination_phone,
                        delivery_national_id=delivery_national_id,
                        order_weight=order_weight,
                        payment_period_1_days=payment_period_1_days,
                        payment_period_2_days=payment_period_2_days,
                        payment_period_3_days=payment_period_3_days,
                        payment_amount_1=payment_amount_1,
                        payment_amount_2=payment_amount_2,
                        payment_amount_3=payment_amount_3,
                        shipped_weight=shipped_weight,
                        unshipped_weight=unshipped_weight
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_rows.append(f"ردیف {row_idx}: خطا در پردازش - {str(e)}")
            
            return success_count, error_rows, old_count
            
        except Exception as e:
            return 0, [f"خطا در خواندن فایل: {str(e)}"], 0
    
    def _safe_decimal(self, value):
        """Safely convert value to Decimal"""
        if value is None or value == '':
            return Decimal('0')
        try:
            cleaned = self.clean_numeric_string(value)
            return Decimal(str(cleaned))
        except:
            return Decimal('0')
    
    def _safe_integer(self, value):
        """Safely convert value to integer"""
        if value is None or value == '':
            return None
        try:
            return int(float(str(value)))
        except:
            return None
    
    def _safe_boolean(self, value):
        """Safely convert value to boolean"""
        if value is None:
            return False
        value_str = str(value).strip().lower()
        return value_str in ['true', '1', 'yes', 'بله', 'درست']
    
    def _map_user_type(self, value):
        """Map Persian user type to English"""
        mapping = {
            'حقیقی': 'individual',
            'حقوقی': 'company'
        }
        return mapping.get(value, 'individual')


# View functions
@staff_member_required
def download_delivery_template(request):
    """دانلود نمونه فایل اکسل برای آدرس‌های تحویل"""
    processor = DeliveryAddressProcessor()
    workbook = processor.create_delivery_template()
    filename = "delivery_addresses_template.xlsx"
    return processor.create_excel_response(workbook, filename)


@staff_member_required
def upload_delivery_addresses(request, purchase_detail_id):
    """آپلود آدرس‌های تحویل با پردازش بهبود یافته"""
    purchase_detail = get_object_or_404(MarketplacePurchaseDetail, id=purchase_detail_id)
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        processor = DeliveryAddressProcessor()
        success_count, error_rows, old_count = processor.process_delivery_upload(
            request.FILES['excel_file'], purchase_detail
        )
        
        if old_count > 0:
            messages.info(request, f'🔄 {old_count} آدرس قبلی حذف شد')
        
        if error_rows:
            for error in error_rows:
                messages.error(request, error)
        
        if success_count > 0:
            messages.success(request, f'✅ تعداد {success_count} آدرس تحویل با موفقیت ثبت شد')
        else:
            messages.warning(request, '⚠️ هیچ آدرسی ثبت نشد')
        
        # Redirect back to admin after successful upload
        return redirect(f'/admin/marketplace/marketplacepurchasedetail/{purchase_detail_id}/change/')
    
    # Show upload form for GET requests
    return render(request, 'marketplace/upload_delivery_addresses.html', {
        'purchase_detail': purchase_detail,
        'title': 'آپلود آدرس‌های تحویل'
    })