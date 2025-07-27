from django.shortcuts import render
from django.db.models import Sum, Count, Q
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from django.http import JsonResponse, HttpResponse
from django_jalali.db.models import jDateField
from django.contrib import messages
from django.shortcuts import redirect, get_object_or_404
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
from .models import (
    WarehouseInventory, PurchaseProforma, SalesProforma, 
    WarehouseReceipt, ProductDelivery, Warehouse, Product, 
    ProductCategory, Supplier, Customer, WarehouseDeliveryOrder,
    WarehouseDeliveryOrderItem, Receiver
)
import jdatetime

class BaseReportView(TemplateView):
    """کلاس پایه برای گزارش‌ها"""
    
    @method_decorator(staff_member_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def format_number(self, value):
        """فرمت کردن اعداد با جداکننده هزارگان"""
        if value is None:
            return 0
        return f'{int(value):,}'

class WarehouseReportView(BaseReportView):
    """گزارش مرور انبار"""
    template_name = 'warehouse/reports/warehouse_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # فیلترهای جستجو
        warehouse_id = self.request.GET.get('warehouse')
        category_id = self.request.GET.get('category')
        product_id = self.request.GET.get('product')
        
        # موجودی‌های انبار
        inventories = WarehouseInventory.objects.select_related('warehouse', 'product', 'product__category')
        
        if warehouse_id:
            inventories = inventories.filter(warehouse_id=warehouse_id)
        if category_id:
            inventories = inventories.filter(product__category_id=category_id)
        if product_id:
            inventories = inventories.filter(product_id=product_id)
        
        # محاسبه آمار کلی
        total_products = inventories.count()
        total_quantity = inventories.aggregate(Sum('quantity'))['quantity__sum'] or 0
        total_reserved = inventories.aggregate(Sum('reserved_quantity'))['reserved_quantity__sum'] or 0
        total_available = inventories.aggregate(Sum('available_quantity'))['available_quantity__sum'] or 0
        
        # گروه‌بندی بر اساس انبار
        warehouse_summary = {}
        for inventory in inventories:
            warehouse_name = inventory.warehouse.name
            if warehouse_name not in warehouse_summary:
                warehouse_summary[warehouse_name] = {
                    'total_quantity': 0,
                    'total_reserved': 0,
                    'total_available': 0,
                    'product_count': 0
                }
            warehouse_summary[warehouse_name]['total_quantity'] += inventory.quantity
            warehouse_summary[warehouse_name]['total_reserved'] += inventory.reserved_quantity
            warehouse_summary[warehouse_name]['total_available'] += inventory.available_quantity
            warehouse_summary[warehouse_name]['product_count'] += 1
        
        context.update({
            'inventories': inventories.order_by('warehouse__name', 'product__name'),
            'warehouses': Warehouse.objects.all(),
            'categories': ProductCategory.objects.all(),
            'products': Product.objects.all(),
            'total_products': total_products,
            'total_quantity': self.format_number(total_quantity),
            'total_reserved': self.format_number(total_reserved),
            'total_available': self.format_number(total_available),
            'warehouse_summary': warehouse_summary,
            'selected_warehouse': warehouse_id,
            'selected_category': category_id,
            'selected_product': product_id,
        })
        
        return context

class PurchaseReportView(BaseReportView):
    """گزارش مرور خرید"""
    template_name = 'warehouse/reports/purchase_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # فیلترهای جستجو
        supplier_id = self.request.GET.get('supplier')
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')
        
        # پیش‌فاکتورهای خرید
        purchase_proformas = PurchaseProforma.objects.select_related('supplier')
        
        if supplier_id:
            purchase_proformas = purchase_proformas.filter(supplier_id=supplier_id)
        if from_date:
            # تبدیل تاریخ شمسی به میلادی
            try:
                from_date_jalali = jdatetime.datetime.strptime(from_date, '%Y/%m/%d').date()
                from_date_gregorian = from_date_jalali.togregorian()
                purchase_proformas = purchase_proformas.filter(date__gte=from_date_gregorian)
            except:
                pass
        if to_date:
            try:
                to_date_jalali = jdatetime.datetime.strptime(to_date, '%Y/%m/%d').date()
                to_date_gregorian = to_date_jalali.togregorian()
                purchase_proformas = purchase_proformas.filter(date__lte=to_date_gregorian)
            except:
                pass
        
        # محاسبه آمار کلی
        total_proformas = purchase_proformas.count()
        total_amount = purchase_proformas.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        # رسیدهای انبار مرتبط
        receipts = WarehouseReceipt.objects.filter(
            purchase_proforma__in=purchase_proformas
        ).select_related('warehouse', 'purchase_proforma')
        
        total_receipts = receipts.count()
        total_received_weight = receipts.aggregate(Sum('total_weight'))['total_weight__sum'] or 0
        
        # گروه‌بندی بر اساس تامین‌کننده
        supplier_summary = {}
        for proforma in purchase_proformas:
            supplier_name = str(proforma.supplier)
            if supplier_name not in supplier_summary:
                supplier_summary[supplier_name] = {
                    'proforma_count': 0,
                    'total_amount': 0,
                    'receipt_count': 0
                }
            supplier_summary[supplier_name]['proforma_count'] += 1
            supplier_summary[supplier_name]['total_amount'] += proforma.total_amount
            supplier_summary[supplier_name]['receipt_count'] = receipts.filter(
                purchase_proforma__supplier=proforma.supplier
            ).count()
        
        # آمار ماهانه
        monthly_stats = {}
        for proforma in purchase_proformas:
            month_key = proforma.date.strftime('%Y-%m')
            if month_key not in monthly_stats:
                monthly_stats[month_key] = {'count': 0, 'amount': 0}
            monthly_stats[month_key]['count'] += 1
            monthly_stats[month_key]['amount'] += proforma.total_amount
        
        context.update({
            'purchase_proformas': purchase_proformas.order_by('-date'),
            'receipts': receipts.order_by('-date'),
            'suppliers': Supplier.objects.all(),
            'total_proformas': total_proformas,
            'total_amount': self.format_number(total_amount),
            'total_receipts': total_receipts,
            'total_received_weight': self.format_number(total_received_weight),
            'supplier_summary': supplier_summary,
            'monthly_stats': monthly_stats,
            'selected_supplier': supplier_id,
            'selected_from_date': from_date,
            'selected_to_date': to_date,
        })
        
        return context

class SalesReportView(BaseReportView):
    """گزارش مرور فروش"""
    template_name = 'warehouse/reports/sales_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # فیلترهای جستجو
        customer_id = self.request.GET.get('customer')
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')
        
        # پیش‌فاکتورهای فروش
        sales_proformas = SalesProforma.objects.select_related('customer')
        
        if customer_id:
            sales_proformas = sales_proformas.filter(customer_id=customer_id)
        if from_date:
            try:
                from_date_jalali = jdatetime.datetime.strptime(from_date, '%Y/%m/%d').date()
                from_date_gregorian = from_date_jalali.togregorian()
                sales_proformas = sales_proformas.filter(date__gte=from_date_gregorian)
            except:
                pass
        if to_date:
            try:
                to_date_jalali = jdatetime.datetime.strptime(to_date, '%Y/%m/%d').date()
                to_date_gregorian = to_date_jalali.togregorian()
                sales_proformas = sales_proformas.filter(date__lte=to_date_gregorian)
            except:
                pass
        
        # محاسبه آمار کلی
        total_proformas = sales_proformas.count()
        total_amount = sales_proformas.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        # تحویل‌های کالا مرتبط
        from .models import WarehouseDeliveryOrder
        deliveries = ProductDelivery.objects.filter(
            delivery_order__sales_proforma__in=sales_proformas
        ).select_related('exit_warehouse', 'delivery_order')
        
        total_deliveries = deliveries.count()
        total_delivered_weight = deliveries.aggregate(Sum('total_weight'))['total_weight__sum'] or 0
        
        # گروه‌بندی بر اساس مشتری
        customer_summary = {}
        for proforma in sales_proformas:
            customer_name = str(proforma.customer)
            if customer_name not in customer_summary:
                customer_summary[customer_name] = {
                    'proforma_count': 0,
                    'total_amount': 0,
                    'delivery_count': 0
                }
            customer_summary[customer_name]['proforma_count'] += 1
            customer_summary[customer_name]['total_amount'] += proforma.total_amount
            customer_summary[customer_name]['delivery_count'] = deliveries.filter(
                delivery_order__sales_proforma__customer=proforma.customer
            ).count()
        
        # آمار ماهانه
        monthly_stats = {}
        for proforma in sales_proformas:
            month_key = proforma.date.strftime('%Y-%m')
            if month_key not in monthly_stats:
                monthly_stats[month_key] = {'count': 0, 'amount': 0}
            monthly_stats[month_key]['count'] += 1
            monthly_stats[month_key]['amount'] += proforma.total_amount
        
        context.update({
            'sales_proformas': sales_proformas.order_by('-date'),
            'deliveries': deliveries.order_by('-exit_date'),
            'customers': Customer.objects.all(),
            'total_proformas': total_proformas,
            'total_amount': self.format_number(total_amount),
            'total_deliveries': total_deliveries,
            'total_delivered_weight': self.format_number(total_delivered_weight),
            'customer_summary': customer_summary,
            'monthly_stats': monthly_stats,
            'selected_customer': customer_id,
            'selected_from_date': from_date,
            'selected_to_date': to_date,
        })
        
        return context

class DashboardView(BaseReportView):
    """داشبورد اصلی"""
    template_name = 'warehouse/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # آمار کلی
        total_warehouses = Warehouse.objects.count()
        total_products = Product.objects.count()
        total_suppliers = Supplier.objects.count()
        total_customers = Customer.objects.count()
        
        # آمار مالی
        total_purchase_amount = PurchaseProforma.objects.aggregate(
            Sum('total_amount'))['total_amount__sum'] or 0
        total_sales_amount = SalesProforma.objects.aggregate(
            Sum('total_amount'))['total_amount__sum'] or 0
        
        # آمار انبار
        total_inventory = WarehouseInventory.objects.aggregate(
            Sum('quantity'))['quantity__sum'] or 0
        total_reserved = WarehouseInventory.objects.aggregate(
            Sum('reserved_quantity'))['reserved_quantity__sum'] or 0
        
        # آخرین فعالیت‌ها
        recent_purchases = PurchaseProforma.objects.order_by('-created_at')[:5]
        recent_sales = SalesProforma.objects.order_by('-created_at')[:5]
        recent_receipts = WarehouseReceipt.objects.order_by('-created_at')[:5]
        recent_deliveries = ProductDelivery.objects.order_by('-created_at')[:5]
        
        context.update({
            'total_warehouses': total_warehouses,
            'total_products': total_products,
            'total_suppliers': total_suppliers,
            'total_customers': total_customers,
            'total_purchase_amount': self.format_number(total_purchase_amount),
            'total_sales_amount': self.format_number(total_sales_amount),
            'total_inventory': self.format_number(total_inventory),
            'total_reserved': self.format_number(total_reserved),
            'recent_purchases': recent_purchases,
            'recent_sales': recent_sales,
            'recent_receipts': recent_receipts,
            'recent_deliveries': recent_deliveries,
        })
        
        return context

# APIهای AJAX برای چارت‌ها
@staff_member_required
def warehouse_chart_data(request):
    """داده‌های چارت موجودی انبار"""
    warehouses = Warehouse.objects.all()
    data = []
    
    for warehouse in warehouses:
        inventory = WarehouseInventory.objects.filter(warehouse=warehouse).aggregate(
            total=Sum('quantity'))['total'] or 0
        data.append({
            'name': warehouse.name,
            'value': float(inventory)
        })
    
    return JsonResponse(data, safe=False)

@staff_member_required
def upload_delivery_order_excel(request, delivery_order_id):
    """آپلود فایل اکسل برای آیتم‌های حواله خروج"""
    delivery_order = get_object_or_404(WarehouseDeliveryOrder, id=delivery_order_id)
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            # خواندن فایل اکسل
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            success_count = 0
            error_rows = []
            
            # خواندن داده‌ها از ردیف دوم (ردیف اول هدر)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # فرض کنیم ستون‌ها اینطوری هستن:
                    # A: کد کالا, B: مقدار, C: نوع وسیله, D: شناسه گیرنده
                    if not any(row):  # ردیف خالی
                        continue
                        
                    product_code = row[0]
                    quantity = row[1] 
                    vehicle_type = row[2]
                    receiver_unique_id = row[3]
                    
                    # پیدا کردن کالا
                    product = Product.objects.get(code=product_code)
                    
                    # پیدا کردن گیرنده
                    receiver = Receiver.objects.get(unique_id=receiver_unique_id)
                    
                    # ایجاد آیتم حواله
                    item = WarehouseDeliveryOrderItem.objects.create(
                        delivery_order=delivery_order,
                        product=product,
                        quantity=quantity,
                        vehicle_type=vehicle_type,
                        receiver=receiver,
                        receiver_address=receiver.address,
                        receiver_postal_code=receiver.postal_code,
                        receiver_phone=receiver.phone,
                        receiver_unique_id=receiver.unique_id
                    )
                    success_count += 1
                    
                except Exception as e:
                    error_rows.append(f"ردیف {row_idx}: {str(e)}")
            
            if success_count > 0:
                messages.success(request, f'{success_count} آیتم با موفقیت اضافه شد')
            
            if error_rows:
                messages.warning(request, f'خطا در ردیف‌های: {", ".join(error_rows)}')
                
        except Exception as e:
            messages.error(request, f'خطا در خواندن فایل: {str(e)}')
    
    return redirect('admin:warehouse_warehousedeliveryorder_change', delivery_order_id)

@staff_member_required
def download_delivery_order_excel(request, delivery_order_id):
    """دانلود آیتم‌های حواله خروج به فرمت اکسل"""
    delivery_order = get_object_or_404(WarehouseDeliveryOrder, id=delivery_order_id)
    
    # ایجاد فایل اکسل
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"حواله {delivery_order.number}"
    
    # تنظیمات استایل
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # هدرها
    headers = [
        'ردیف', 'کالا', 'کد کالا', 'مقدار', 'واحد', 'نوع وسیله', 
        'گیرنده', 'شناسه گیرنده', 'تلفن گیرنده', 'آدرس گیرنده', 'کد پستی'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # داده‌های آیتم‌ها
    items = delivery_order.items.all().order_by('row_number')
    for row_idx, item in enumerate(items, 2):
        sheet.cell(row=row_idx, column=1, value=item.row_number)
        sheet.cell(row=row_idx, column=2, value=item.product.name)
        sheet.cell(row=row_idx, column=3, value=item.product.code)
        sheet.cell(row=row_idx, column=4, value=float(item.quantity))
        sheet.cell(row=row_idx, column=5, value=item.product.unit)
        sheet.cell(row=row_idx, column=6, value=item.get_vehicle_type_display())
        sheet.cell(row=row_idx, column=7, value=str(item.receiver))
        sheet.cell(row=row_idx, column=8, value=item.receiver_unique_id)
        sheet.cell(row=row_idx, column=9, value=item.receiver_phone)
        sheet.cell(row=row_idx, column=10, value=item.receiver_address)
        sheet.cell(row=row_idx, column=11, value=item.receiver_postal_code)
    
    # تنظیم عرض ستون‌ها
    column_widths = [8, 25, 15, 12, 10, 15, 25, 15, 15, 40, 12]
    for col_idx, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # ایجاد response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="delivery_order_{delivery_order.number}.xlsx"'
    
    # نوشتن فایل در response
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response.write(virtual_workbook.getvalue())
    
    return response

@staff_member_required
def upload_delivery_order_page(request, delivery_order_id):
    """صفحه آپلود فایل اکسل"""
    delivery_order = get_object_or_404(WarehouseDeliveryOrder, id=delivery_order_id)
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            # خواندن فایل اکسل
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            success_count = 0
            error_rows = []
            
            # خواندن داده‌ها از ردیف دوم (ردیف اول هدر)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not any(row):  # ردیف خالی
                        continue
                        
                    product_code = row[0]
                    quantity = row[1] 
                    vehicle_type = row[2]
                    receiver_unique_id = row[3]
                    
                    if not all([product_code, quantity, vehicle_type, receiver_unique_id]):
                        error_rows.append(f"ردیف {row_idx}: داده‌های ناقص")
                        continue
                    
                    # پیدا کردن کالا
                    product = Product.objects.get(code=product_code)
                    
                    # پیدا کردن گیرنده
                    receiver = Receiver.objects.get(unique_id=receiver_unique_id)
                    
                    # ایجاد آیتم حواله
                    item = WarehouseDeliveryOrderItem.objects.create(
                        delivery_order=delivery_order,
                        product=product,
                        quantity=quantity,
                        vehicle_type=vehicle_type,
                        receiver=receiver,
                        receiver_address=receiver.address,
                        receiver_postal_code=receiver.postal_code,
                        receiver_phone=receiver.phone,
                        receiver_unique_id=receiver.unique_id
                    )
                    success_count += 1
                    
                except Product.DoesNotExist:
                    error_rows.append(f"ردیف {row_idx}: کالا با کد {product_code} یافت نشد")
                except Receiver.DoesNotExist:
                    error_rows.append(f"ردیف {row_idx}: گیرنده با شناسه {receiver_unique_id} یافت نشد")
                except Exception as e:
                    error_rows.append(f"ردیف {row_idx}: {str(e)}")
            
            if success_count > 0:
                messages.success(request, f'{success_count} آیتم با موفقیت اضافه شد')
            
            if error_rows:
                error_message = f'خطا در ردیف‌های: ' + ', '.join(error_rows[:5])
                if len(error_rows) > 5:
                    error_message += f' و {len(error_rows) - 5} خطای دیگر'
                messages.warning(request, error_message)
                
        except Exception as e:
            messages.error(request, f'خطا در خواندن فایل: {str(e)}')
        
        return redirect(f'/admin/warehouse/warehousedeliveryorder/{delivery_order_id}/change/')
    
    context = {
        'delivery_order': delivery_order,
        'title': f'آپلود فایل اکسل برای حواله {delivery_order.number}'
    }
    return render(request, 'warehouse/upload_excel.html', context)

@staff_member_required
def download_delivery_order_template(request):
    """دانلود فایل نمونه برای آپلود حواله"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "نمونه حواله"
    
    # تنظیمات استایل
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # هدرها
    headers = ['کد کالا', 'مقدار', 'نوع وسیله حمل', 'شناسه گیرنده']
    descriptions = [
        'کد کالای تعریف شده در سیستم',
        'مقدار کالا (عدد)',
        'truck/pickup/van/container/other',
        'شناسه یکتای گیرنده'
    ]
    
    for col_idx, (header, desc) in enumerate(zip(headers, descriptions), 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        
        # توضیحات در ردیف دوم
        desc_cell = sheet.cell(row=2, column=col_idx, value=desc)
        desc_cell.font = Font(italic=True, color="666666")
    
    # نمونه داده
    sample_data = [
        ['K001', 100, 'truck', 'R001'],
        ['K002', 50, 'pickup', 'R002'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # تنظیم عرض ستون‌ها
    for col_idx in range(1, 5):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25
    
    # ایجاد response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="delivery_order_template.xlsx"'
    
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response.write(virtual_workbook.getvalue())
    
    return response

@staff_member_required
def monthly_sales_chart_data(request):
    """داده‌های چارت فروش ماهانه"""
    import datetime
    from django.db.models import Extract
    
    # آمار 12 ماه گذشته
    monthly_data = SalesProforma.objects.filter(
        date__gte=datetime.date.today() - datetime.timedelta(days=365)
    ).extra(
        select={'month': "EXTRACT(month FROM date)", 'year': "EXTRACT(year FROM date)"}
    ).values('month', 'year').annotate(
        total=Sum('total_amount'), count=Count('id')
    ).order_by('year', 'month')
    
    data = []
    for item in monthly_data:
        data.append({
            'month': f"{item['year']}-{item['month']:02d}",
            'amount': float(item['total']),
            'count': item['count']
        })
    
    return JsonResponse(data, safe=False)