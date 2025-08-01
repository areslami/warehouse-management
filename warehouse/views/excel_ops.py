from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
from ..models import (
    WarehouseDeliveryOrder, WarehouseDeliveryOrderItem, 
    Product, Receiver
)

@staff_member_required
def get_receiver_info(request, receiver_id):
    """API برای دریافت اطلاعات گیرنده"""
    try:
        receiver = get_object_or_404(Receiver, id=receiver_id)
        data = {
            'address': receiver.address,
            'phone': receiver.phone,
            'postal_code': receiver.postal_code,
            'unique_id': receiver.unique_id,
        }
        return JsonResponse(data)
    except:
        return JsonResponse({'error': 'گیرنده یافت نشد'}, status=404)

@staff_member_required
def upload_delivery_order_excel(request, delivery_order_id):
    """آپلود فایل اکسل برای آیتم‌های حواله خروج"""
    delivery_order = get_object_or_404(WarehouseDeliveryOrder, id=delivery_order_id)
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            # خواندن فایل اکسل
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            success_count = 0
            error_rows = []
            
            # خواندن داده‌ها از ردیف دوم (ردیف اول هدر)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not any(row):  # ردیف خالی
                        continue
                        
                    product_code = str(row[0]).strip() if row[0] else None
                    quantity = float(row[1]) if row[1] else None
                    vehicle_type = str(row[2]).strip().lower() if row[2] else None
                    receiver_unique_id = str(row[3]).strip() if row[3] else None
                    
                    if not all([product_code, quantity, vehicle_type, receiver_unique_id]):
                        error_rows.append(f"ردیف {row_idx}: داده‌های ناقص")
                        continue
                    
                    # بررسی نوع وسیله حمل معتبر
                    valid_vehicles = ['truck', 'pickup', 'van', 'container', 'other']
                    if vehicle_type not in valid_vehicles:
                        error_rows.append(f"ردیف {row_idx}: نوع وسیله نامعتبر ({vehicle_type})")
                        continue
                    
                    # پیدا کردن کالا
                    product = Product.objects.get(code=product_code)
                    
                    # پیدا کردن گیرنده
                    receiver = Receiver.objects.get(unique_id=receiver_unique_id)
                    
                    # ایجاد آیتم حواله
                    item = WarehouseDeliveryOrderItem.objects.create(
                        delivery_order=delivery_order,
                        product=product,
                        quantity=quantity,
                        vehicle_type=vehicle_type,
                        receiver=receiver,
                        receiver_address=receiver.address,
                        receiver_postal_code=receiver.postal_code,
                        receiver_phone=receiver.phone,
                        receiver_unique_id=receiver.unique_id
                    )
                    success_count += 1
                    
                except Product.DoesNotExist:
                    error_rows.append(f"ردیف {row_idx}: کالا با کد '{product_code}' یافت نشد")
                except Receiver.DoesNotExist:
                    error_rows.append(f"ردیف {row_idx}: گیرنده با شناسه '{receiver_unique_id}' یافت نشد")
                except ValueError:
                    error_rows.append(f"ردیف {row_idx}: مقدار نامعتبر")
                except Exception as e:
                    error_rows.append(f"ردیف {row_idx}: {str(e)}")
            
            if success_count > 0:
                messages.success(request, f'✅ {success_count} آیتم با موفقیت اضافه شد')
            
            if error_rows:
                error_message = '❌ خطا در ردیف‌های: ' + ', '.join(error_rows[:5])
                if len(error_rows) > 5:
                    error_message += f' و {len(error_rows) - 5} خطای دیگر'
                messages.warning(request, error_message)
                
        except Exception as e:
            messages.error(request, f'❌ خطا در خواندن فایل: {str(e)}')
        
        return redirect(f'/admin/warehouse/warehousedeliveryorder/{delivery_order_id}/change/')
    
    context = {
        'delivery_order': delivery_order,
        'title': f'آپلود فایل اکسل برای حواله {delivery_order.number}'
    }
    return render(request, 'warehouse/upload_excel.html', context)

@staff_member_required
def download_delivery_order_excel(request, delivery_order_id):
    """دانلود آیتم‌های حواله خروج به فرمت اکسل"""
    delivery_order = get_object_or_404(WarehouseDeliveryOrder, id=delivery_order_id)
    
    # ایجاد فایل اکسل
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"حواله {delivery_order.number}"
    
    # تنظیمات استایل
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # هدرها
    headers = [
        'ردیف', 'کالا', 'کد کالا', 'مقدار', 'واحد', 'نوع وسیله', 
        'گیرنده', 'شناسه گیرنده', 'تلفن گیرنده', 'آدرس گیرنده', 'کد پستی'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # داده‌های آیتم‌ها
    items = delivery_order.items.all().order_by('row_number')
    for row_idx, item in enumerate(items, 2):
        sheet.cell(row=row_idx, column=1, value=item.row_number)
        sheet.cell(row=row_idx, column=2, value=item.product.name)
        sheet.cell(row=row_idx, column=3, value=item.product.code)
        sheet.cell(row=row_idx, column=4, value=float(item.quantity))
        sheet.cell(row=row_idx, column=5, value=item.product.unit)
        sheet.cell(row=row_idx, column=6, value=item.get_vehicle_type_display())
        sheet.cell(row=row_idx, column=7, value=str(item.receiver))
        sheet.cell(row=row_idx, column=8, value=item.receiver_unique_id)
        sheet.cell(row=row_idx, column=9, value=item.receiver_phone)
        sheet.cell(row=row_idx, column=10, value=item.receiver_address)
        sheet.cell(row=row_idx, column=11, value=item.receiver_postal_code)
    
    # تنظیم عرض ستون‌ها
    column_widths = [8, 25, 15, 12, 10, 15, 25, 15, 15, 40, 12]
    for col_idx, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # ایجاد response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="delivery_order_{delivery_order.number}.xlsx"'
    
    # نوشتن فایل در response
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response.write(virtual_workbook.getvalue())
    
    return response

@staff_member_required
def download_delivery_order_template(request):
    """دانلود فایل نمونه برای آپلود حواله"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "نمونه حواله"
    
    # تنظیمات استایل
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # هدرها
    headers = ['کد کالا', 'مقدار', 'نوع وسیله حمل', 'شناسه گیرنده']
    descriptions = [
        'کد کالای تعریف شده در سیستم',
        'مقدار کالا (عدد)',
        'truck/pickup/van/container/other',
        'شناسه یکتای گیرنده'
    ]
    
    for col_idx, (header, desc) in enumerate(zip(headers, descriptions), 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        
        # توضیحات در ردیف دوم
        desc_cell = sheet.cell(row=2, column=col_idx, value=desc)
        desc_cell.font = Font(italic=True, color="666666")
    
    # نمونه داده
    sample_data = [
        ['K001', 100, 'truck', 'R001'],
        ['K002', 50, 'pickup', 'R002'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # تنظیم عرض ستون‌ها
    for col_idx in range(1, 5):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25
    
    # ایجاد response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="delivery_order_template.xlsx"'
    
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    response.write(virtual_workbook.getvalue())
    
    return response


@staff_member_required
def bulk_delivery_order_selection(request):
    """صفحه انتخاب حواله‌های خروج برای ارسال بالک"""
    from django.db.models import Q
    from django.core.paginator import Paginator
    
    # دریافت پارامترهای فیلتر
    search_query = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    warehouse_id = request.GET.get('warehouse', '')
    
    # شروع با کل حواله‌ها
    delivery_orders = WarehouseDeliveryOrder.objects.select_related(
        'warehouse', 'sales_proforma', 'shipping_company'
    ).prefetch_related('items')
    
    # اعمال فیلترها
    if search_query:
        delivery_orders = delivery_orders.filter(
            Q(number__icontains=search_query) |
            Q(sales_proforma__number__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    if date_from:
        delivery_orders = delivery_orders.filter(issue_date__gte=date_from)
    
    if date_to:
        delivery_orders = delivery_orders.filter(issue_date__lte=date_to)
    
    if warehouse_id:
        delivery_orders = delivery_orders.filter(warehouse_id=warehouse_id)
    
    # مرتب‌سازی
    delivery_orders = delivery_orders.order_by('-issue_date', '-number')
    
    # صفحه‌بندی
    paginator = Paginator(delivery_orders, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # دریافت لیست انبارها برای فیلتر
    from ..models.base import Warehouse
    warehouses = Warehouse.objects.all()
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'warehouse_id': warehouse_id,
        'warehouses': warehouses,
        'total_count': paginator.count
    }
    
    return render(request, 'warehouse/bulk_delivery_selection.html', context)


@staff_member_required
def bulk_delivery_order_export(request):
    """ارسال بالک حواله‌های خروج به اکسل"""
    if request.method != 'POST':
        messages.error(request, 'روش درخواست نامعتبر است')
        return redirect('warehouse:bulk_delivery_order_selection')
    
    # دریافت شناسه‌های حواله انتخاب شده
    delivery_order_ids = request.POST.getlist('delivery_orders')
    
    if not delivery_order_ids:
        messages.error(request, 'هیچ حواله‌ای انتخاب نشده است')
        return redirect('warehouse:bulk_delivery_order_selection')
    
    try:
        # دریافت حواله‌های انتخاب شده
        delivery_orders = WarehouseDeliveryOrder.objects.filter(
            id__in=delivery_order_ids
        ).select_related(
            'warehouse', 'sales_proforma', 'shipping_company'
        ).prefetch_related(
            'items__product', 'items__receiver'
        )
        
        if not delivery_orders.exists():
            messages.error(request, 'حواله‌های انتخاب شده یافت نشد')
            return redirect('warehouse:bulk_delivery_order_selection')
        
        # ایجاد فایل اکسل
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "حواله‌های خروج انبار"
        
        # تنظیم سبک‌ها
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # هدرهای جدول
        headers = [
            'شماره حواله', 'تاریخ صدور', 'انبار', 'پیش‌فاکتور فروش', 'شرکت حمل',
            'کد کالا', 'نام کالا', 'مقدار', 'واحد', 'نوع وسیله حمل',
            'کد گیرنده', 'نام گیرنده', 'آدرس گیرنده', 'تلفن گیرنده', 'کد پستی گیرنده', 'شناسه یکتای گیرنده'
        ]
        
        # نوشتن هدرها
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # نوشتن داده‌ها
        row_idx = 2
        
        for delivery_order in delivery_orders:
            for item in delivery_order.items.all():
                data_row = [
                    delivery_order.number,
                    delivery_order.issue_date.strftime('%Y/%m/%d') if delivery_order.issue_date else '',
                    delivery_order.warehouse.name if delivery_order.warehouse else '',
                    delivery_order.sales_proforma.number if delivery_order.sales_proforma else '',
                    delivery_order.shipping_company.name if delivery_order.shipping_company else '',
                    item.product.code if item.product else '',
                    item.product.name if item.product else '',
                    str(item.quantity),
                    item.product.unit if item.product else '',
                    item.get_vehicle_type_display(),
                    item.receiver.code if item.receiver else '',
                    item.receiver.full_name if item.receiver and item.receiver.receiver_type == 'natural' else (item.receiver.company_name if item.receiver else ''),
                    item.receiver_address,
                    item.receiver_phone,
                    item.receiver_postal_code,
                    item.receiver_unique_id
                ]
                
                for col_idx, value in enumerate(data_row, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
                
                row_idx += 1
        
        # تنظیم عرض ستون‌ها
        for col_idx in range(1, len(headers) + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
        
        # ایجاد response برای دانلود
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # نام فایل با تاریخ
        import jdatetime
        today = jdatetime.date.today()
        filename = f"bulk_delivery_orders_{today.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # ذخیره فایل
        virtual_workbook = io.BytesIO()
        workbook.save(virtual_workbook)
        virtual_workbook.seek(0)
        response.write(virtual_workbook.getvalue())
        
        # پیام موفقیت
        messages.success(request, f'فایل اکسل حاوی {len(delivery_order_ids)} حواله با موفقیت ایجاد شد')
        
        return response
        
    except Exception as e:
        messages.error(request, f'خطا در ایجاد فایل اکسل: {str(e)}')
        return redirect('warehouse:bulk_delivery_order_selection')